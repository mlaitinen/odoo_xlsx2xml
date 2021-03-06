import os
import sys
from openpyxl import load_workbook
from lxml import etree


def generate_accounts(sheet, data, record_model):

    # Get the header values / field names
    id_col = None
    headers = {}
    attrs = {}
    for header_cell in sheet.rows.__next__():
        header_value = header_cell.value
        if not header_value:
            continue

        if header_value == 'id':
            id_col = header_cell.column
            continue

        if '?' in header_value:
            header_value, attr = header_value.split('?')
            attrs[header_cell.column] = attr

        headers[header_cell.column] = header_value

    if id_col is None:
        raise Exception("The ID column is missing")

    def create_account(account_row):
        # Create a single account record.
        record = etree.Element("record", model=record_model)

        for cell in account_row:
            value = str(cell.value)

            if cell.column in headers and cell.value:
                field = etree.Element("field", name=headers[cell.column])

                if cell.column in attrs:
                    # As an attribute
                    field.attrib[attrs[cell.column]] = value
                else:
                    # As a value
                    field.text = value

                record.append(field)
            elif cell.column == id_col:
                record.attrib['id'] = value

        return record

    # Iterate the rows in the worksheet
    for row in sheet.iter_rows(row_offset=1):
        data.append(create_account(row))


def main():
    # Get parameters
    if len(sys.argv) < 3:
        print("Usage: {} <xlsx file> <record model> [<output.xml>]".format(os.path.basename(__file__)))
        exit()
    elif len(sys.argv) >= 4:
        output_file = sys.argv[3]
    else:
        output_file = 'output.xml'

    src_filename = sys.argv[1]

    try:
        sheet = load_workbook(filename=src_filename, use_iterators=True).worksheets[0]
    except IOError:
        print("No such file:", src_filename)
        exit()

    # Create the XML tree
    root = etree.Element("openerp")
    data = etree.Element("data", noupdate="1")
    root.append(data)
    generate_accounts(sheet, data, sys.argv[2])

    # Write the XML content to the output file
    xml_str = etree.tostring(root, pretty_print=True)
    with open(output_file, 'wb') as f:
        f.write(xml_str)

if __name__ == "__main__":
    main()